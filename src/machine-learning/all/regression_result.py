import os, logging, sys, json, csv
sys.path.insert(0, '../')

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

from util.util import getScore, all_day_pct_change_negative, all_day_pct_change_positive, no_doji_or_spinning_buy_india, no_doji_or_spinning_sell_india, scrip_patterns_to_dict
from util.util import is_algo_buy, is_algo_sell
from util.util import get_regressionResult
from util.util import buy_pattern_from_history, buy_all_rule, buy_year_high, buy_year_low, buy_up_trend, buy_down_trend, buy_final, buy_high_indicators, buy_pattern
from util.util import sell_pattern_from_history, sell_all_rule, sell_year_high, sell_year_low, sell_up_trend, sell_down_trend, sell_final, sell_high_indicators, sell_pattern
from util.util import buy_pattern_without_mlalgo, sell_pattern_without_mlalgo, buy_oi, sell_oi, all_withoutml
from util.util import morning_star_sell, buy_oi_candidate, morning_star_buy, sell_oi_candidate
from util.util import buy_all_filter, buy_all_common, sell_all_filter, sell_all_common
from util.util import buy_all_rule_classifier, sell_all_rule_classifier
from util.util import is_algo_buy_classifier, is_algo_sell_classifier



connection = MongoClient('localhost', 27017)
db = connection.Nsedata

buyPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-buy.csv')
sellPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-sell.csv')

directory = '../../output/final'
logname = '../../output/final' + '/all-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyOI = wb.create_sheet("buyOI")
ws_buyOI.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_buyOIReg = wb.create_sheet("buyOIReg")
ws_buyOIReg.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_buyOICla = wb.create_sheet("buyOICla")
ws_buyOICla.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])

ws_sellOI = wb.create_sheet("sellOI")
ws_sellOI.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_sellOIReg = wb.create_sheet("sellOIReg")
ws_sellOIReg.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_sellOICla = wb.create_sheet("sellOICla")
ws_sellOICla.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])

ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_buyAllCommon = wb.create_sheet("BuyAllCommon")
ws_buyAllCommon.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_buyAllFilter = wb.create_sheet("BuyAllFilter")
ws_buyAllFilter.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_buyAllReg = wb.create_sheet("BuyAllReg")
ws_buyAllReg.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_buyAllCla = wb.create_sheet("BuyAllCla")
ws_buyAllCla.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])

ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_sellAllCommon = wb.create_sheet("SellAllCommon")
ws_sellAllCommon.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_sellAllFilter = wb.create_sheet("SellAllFilter")
ws_sellAllFilter.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_sellAllReg = wb.create_sheet("SellAllReg")
ws_sellAllReg.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])
ws_sellAllCla = wb.create_sheet("SellAllCla")
ws_sellAllCla.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange", "yLowChange", "seriesTrend", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Filter", "Avg", "Count"])

def saveReports(run_type=None):
    ws_buyAll.append([""])
    ws_buyOI.append([""])
    ws_buyAllCommon.append([""])
    ws_buyAllFilter.append([""])
    ws_buyAllReg.append([""])
    ws_buyOIReg.append([""])
    ws_buyAllCla.append([""])
    ws_buyOICla.append([""])
        
    ws_sellAll.append([""])
    ws_sellOI.append([""])
    ws_sellAllCommon.append([""])
    ws_sellAllFilter.append([""])
    ws_sellAllReg.append([""])
    ws_sellOIReg.append([""])
    ws_sellAllCla.append([""])
    ws_sellOICla.append([""])
    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_buyOI.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyOI.add_table(tab)
    
    count = 0
    for row in ws_buyAllCommon.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllCommon.add_table(tab)
    
    count = 0
    for row in ws_buyAllFilter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllFilter.add_table(tab)
    
    count = 0
    for row in ws_buyAllReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllReg.add_table(tab)
    
    count = 0
    for row in ws_buyOIReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyOIReg.add_table(tab)
    
    count = 0
    for row in ws_buyAllCla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllCla.add_table(tab)
    
    count = 0
    for row in ws_buyOICla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_buyOICla.add_table(tab)
    
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_sellOI.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellOI.add_table(tab)
    
    count = 0
    for row in ws_sellAllCommon.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllCommon.add_table(tab)
    
    count = 0
    for row in ws_sellAllFilter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllFilter.add_table(tab)
    
    count = 0
    for row in ws_sellAllReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllReg.add_table(tab)
    
    count = 0
    for row in ws_sellOIReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellOIReg.add_table(tab)
    
    count = 0
    for row in ws_sellAllCla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllCla.add_table(tab)
    
    count = 0
    for row in ws_sellOICla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AE" + str(count))
    tab.tableStyleInfo = style
    ws_sellOICla.add_table(tab)
    
    
    wb.save(logname + ".xlsx")
      
def result_data(scrip):
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    regressionResult = get_regressionResult(regression_data, scrip, db)
    classificationResult = get_regressionResult(classification_data, scrip, db)
    
    cl_oi = buy_oi_candidate(classification_data, regressionResult, None)
    re_oi = buy_oi_candidate(regression_data, regressionResult, None)
    if(cl_oi and re_oi):
        all_withoutml(regression_data, regressionResult, ws_buyOI)
        all_withoutml(classification_data, classificationResult, ws_buyOI)    
    
    if(regression_data is not None and classification_data is not None
    and (
        (is_algo_buy(regression_data) and is_algo_buy(classification_data))
        or (is_algo_buy(regression_data) and is_algo_buy_classifier(classification_data))
        )
    ):
        buyIndiaAvg, result = buy_pattern_from_history(classification_data, classificationResult, None)
        if (buy_all_rule(regression_data, classificationResult, buyIndiaAvg, None)
            or buy_all_rule_classifier(classification_data, classificationResult, buyIndiaAvg, None)):
            buy_all_filter(classification_data, classificationResult, ws_buyAllFilter)
            buy_all_common(regression_data, classification_data, classificationResult, ws_buyAllCommon)
            if (buy_all_rule(regression_data, classificationResult, buyIndiaAvg, ws_buyAll)
                or buy_all_rule_classifier(classification_data, classificationResult, buyIndiaAvg, ws_buyAll)):
                print('')
            
            buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, None)
            if (buy_all_rule(regression_data, regressionResult, buyIndiaAvg, None)
                or buy_all_rule_classifier(classification_data, regressionResult, buyIndiaAvg, None)):
                buy_all_filter(regression_data, regressionResult, ws_buyAllFilter)
                buy_all_common(regression_data, classification_data, regressionResult, ws_buyAllCommon)
                if (buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAll)
                    or buy_all_rule_classifier(classification_data, regressionResult, buyIndiaAvg, ws_buyAll)):
                    print('')
                 
    
    
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    regressionResult = get_regressionResult(regression_data, scrip, db)
    classificationResult = get_regressionResult(classification_data, scrip, db)
    
    cl_oi = sell_oi_candidate(classification_data, regressionResult, None)
    re_oi = sell_oi_candidate(regression_data, regressionResult, None)
    if(cl_oi and re_oi):
        all_withoutml(regression_data, regressionResult, ws_sellOI)
        all_withoutml(classification_data, classificationResult, ws_sellOI) 
     
    if(regression_data is not None and classification_data is not None
        and (
            (is_algo_sell(regression_data) and is_algo_sell(classification_data))
            or (is_algo_sell(regression_data) and is_algo_sell_classifier(classification_data))
            )
    ):    
        
        sellIndiaAvg, result = sell_pattern_from_history(classification_data, classificationResult, None)
        if (sell_all_rule(regression_data, classificationResult, sellIndiaAvg, None)
            or sell_all_rule_classifier(classification_data, classificationResult, sellIndiaAvg, None)):
            sell_all_filter(classification_data, classificationResult, ws_sellAllFilter)
            sell_all_common(regression_data, classification_data, classificationResult, ws_sellAllCommon)
            if (sell_all_rule(regression_data, classificationResult, sellIndiaAvg, ws_sellAll)
                or sell_all_rule_classifier(classification_data, classificationResult, sellIndiaAvg, ws_sellAll)):
                print('')
    
            sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, None)
            if (sell_all_rule(regression_data, regressionResult, sellIndiaAvg, None)
                or sell_all_rule_classifier(classification_data, regressionResult, sellIndiaAvg, None)):
                sell_all_filter(regression_data, regressionResult, ws_sellAllFilter)
                sell_all_common(regression_data, classification_data, regressionResult, ws_sellAllCommon)
                if (sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAll)
                    or sell_all_rule_classifier(classification_data, regressionResult, sellIndiaAvg, ws_sellAll)):
                    print('')                           

def result_data_reg(scrip):
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db)
        buy_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = buy_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_buyOIReg)
        buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, None)
        if buy_all_rule(regression_data, regressionResult, buyIndiaAvg, None):
            buy_year_high(regression_data, regressionResult, None)
            buy_year_low(regression_data, regressionResult, None, None)
            buy_final(regression_data, regressionResult, None, None)
            buy_high_indicators(regression_data, regressionResult, None)
            buy_up_trend(regression_data, regressionResult, None)
            buy_down_trend(regression_data, regressionResult, None)
            buy_oi(regression_data, regressionResult, None)
            buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAllReg)
                
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db)
        sell_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = sell_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_sellOIReg)
        sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, None)
        if sell_all_rule(regression_data, regressionResult, sellIndiaAvg, None):
            sell_year_high(regression_data, regressionResult, None, None)
            sell_year_low(regression_data, regressionResult, None)
            sell_final(regression_data, regressionResult, None, None)
            sell_high_indicators(regression_data, regressionResult, None)
            sell_up_trend(regression_data, regressionResult, None)
            sell_down_trend(regression_data, regressionResult, None)
            sell_oi(regression_data, regressionResult, None)
            sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAllReg)                                
 
def result_data_cla(scrip):
    regression_data = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db)
        buy_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = buy_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_buyOICla)
        buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, None)
        if buy_all_rule(regression_data, regressionResult, buyIndiaAvg, None):
            buy_year_high(regression_data, regressionResult, None)
            buy_year_low(regression_data, regressionResult, None, None)
            buy_final(regression_data, regressionResult, None, None)
            buy_high_indicators(regression_data, regressionResult, None)
            buy_up_trend(regression_data, regressionResult, None)
            buy_down_trend(regression_data, regressionResult, None)
            buy_oi(regression_data, regressionResult, None)
            buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAllCla)
                        
    regression_data = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db)
        sell_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = sell_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_sellOICla)
        sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, None)
        if sell_all_rule(regression_data, regressionResult, sellIndiaAvg, None):
            sell_year_high(regression_data, regressionResult, None, None)
            sell_year_low(regression_data, regressionResult, None)
            sell_final(regression_data, regressionResult, None, None)
            sell_high_indicators(regression_data, regressionResult, None)
            sell_up_trend(regression_data, regressionResult, None)
            sell_down_trend(regression_data, regressionResult, None)
            sell_oi(regression_data, regressionResult, None)
            sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAllCla)                                 
                                             
def calculateParallel(threads=2, futures=None):
    pool = ThreadPool(threads)
    scrips = []
    for data in db.scrip.find({'futures':futures}):
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    pool.map(result_data, scrips)
    pool.map(result_data_reg, scrips)
    pool.map(result_data_cla, scrips)
                      
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1])
    connection.close()
    saveReports(sys.argv[1])