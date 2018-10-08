import os, logging, sys, json, csv
sys.path.insert(0, '../')

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

from util.util import getScore, all_day_pct_change_negative, all_day_pct_change_positive, no_doji_or_spinning_buy_india, no_doji_or_spinning_sell_india, scrip_patterns_to_dict
from util.util import is_algo_buy, is_algo_sell
from util.util import get_regressionResult
from util.util import buy_pattern_from_history, buy_all_rule, buy_year_high, buy_year_low, buy_up_trend, buy_down_trend, buy_final, buy_high_indicators, buy_pattern
from util.util import sell_pattern_from_history, sell_all_rule, sell_year_high, sell_year_low, sell_up_trend, sell_down_trend, sell_final, sell_high_indicators, sell_pattern
from util.util import buy_pattern_without_mlalgo, sell_pattern_without_mlalgo, buy_oi, sell_oi, all_withoutml
from util.util import buy_oi_candidate, sell_oi_candidate

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

buyPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-buy.csv')
sellPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-sell.csv')

directory = '../../output/final'
logname = '../../output/final' + '/classification-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyYearHigh = wb.create_sheet("buyYearHigh")
ws_buyYearHigh.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyYearLow = wb.create_sheet("buyYearLow")
ws_buyYearLow.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyYearLow1 = wb.create_sheet("buyYearLow1")
ws_buyYearLow1.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyFinal = wb.create_sheet("BuyFinal")
ws_buyFinal.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyFinal1 = wb.create_sheet("BuyFinal1")
ws_buyFinal1.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyHighIndicators = wb.create_sheet("BuyHighIndicators")
ws_buyHighIndicators.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyPattern2 = wb.create_sheet("buyPattern2")
ws_buyPattern2.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyOI = wb.create_sheet("buyOI")
ws_buyOI.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])

ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellYearLow = wb.create_sheet("sellYearLow")
ws_sellYearLow.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellYearHigh = wb.create_sheet("sellYearHigh")
ws_sellYearHigh.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellYearHigh1 = wb.create_sheet("sellYearHigh1")
ws_sellYearHigh1.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellFinal = wb.create_sheet("SellFinal")
ws_sellFinal.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellFinal1 = wb.create_sheet("SellFinal1")
ws_sellFinal1.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellHighIndicators = wb.create_sheet("SellHighIndicators")
ws_sellHighIndicators.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellPattern2 = wb.create_sheet("sellPattern2")
ws_sellPattern2.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellOI = wb.create_sheet("sellOI")
ws_sellOI.append(["BuyIndicators", "SellIndicators", "Symbol", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "trend", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "seriesTrend", "short_term", "long_term", "consolidation", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])

def saveReports(run_type=None):
    ws_buyAll.append([""])
    ws_buyYearHigh.append([""])
    ws_buyYearLow.append([""])
    ws_buyYearLow1.append([""])
    ws_buyFinal.append([""])
    ws_buyFinal1.append([""])
    ws_buyHighIndicators.append([""])
    ws_buyPattern2.append([""])
    ws_buyOI.append([""])
        
    ws_sellAll.append([""])
    ws_sellYearLow.append([""])
    ws_sellYearHigh.append([""])
    ws_sellYearHigh1.append([""])
    ws_sellFinal.append([""])
    ws_sellFinal1.append([""])
    ws_sellHighIndicators.append([""])
    ws_sellPattern2.append([""])
    ws_sellOI.append([""])

    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_buyYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearHigh.add_table(tab)
     
    count = 0
    for row in ws_buyYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearLow.add_table(tab)
   
    count = 0
    for row in ws_buyYearLow1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearLow1.add_table(tab)
     
    count = 0
    for row in ws_buyFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal.add_table(tab)
     
    count = 0
    for row in ws_buyFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal1.add_table(tab)
     
    count = 0
    for row in ws_buyPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern2.add_table(tab)
    
    count = 0
    for row in ws_buyOI.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyOI.add_table(tab)
     
    count = 0
    for row in ws_buyHighIndicators.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_buyHighIndicators.add_table(tab)
    
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_sellYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearHigh.add_table(tab)
    
    count = 0
    for row in ws_sellYearHigh1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearHigh1.add_table(tab)
    
    count = 0
    for row in ws_sellYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearLow.add_table(tab)
    
    count = 0
    for row in ws_sellFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal.add_table(tab)
    
    count = 0
    for row in ws_sellFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal1.add_table(tab)
    
    count = 0
    for row in ws_sellPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern2.add_table(tab)
    
    count = 0
    for row in ws_sellOI.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellOI.add_table(tab)
    
    count = 0
    for row in ws_sellHighIndicators.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AR" + str(count))
    tab.tableStyleInfo = style
    ws_sellHighIndicators.add_table(tab)
    
    wb.save(logname + ".xlsx")
     
def result_data(scrip):
    classification_high = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_low = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})                    
    
    regression_data = classification_high
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, classification_low['mlpValue'], classification_low['kNeighboursValue'])
        buy_pattern_without_mlalgo(regression_data, regressionResult, ws_buyPattern2, ws_sellPattern2)
        buy_oi_candidate(regression_data, regressionResult, None)
        all_withoutml(regression_data, regressionResult, ws_buyOI)
        buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, ws_buyPattern2)
        if buy_all_rule(regression_data, regressionResult, buyIndiaAvg, None):
            buy_year_high(regression_data, regressionResult, ws_buyYearHigh)
            buy_year_low(regression_data, regressionResult, ws_buyYearLow, ws_buyYearLow1)
            buy_final(regression_data, regressionResult, ws_buyFinal, ws_buyFinal1)
            buy_high_indicators(regression_data, regressionResult, ws_buyHighIndicators)
            buy_up_trend(regression_data, regressionResult, None)
            buy_down_trend(regression_data, regressionResult, None)
            buy_oi(regression_data, regressionResult, None)
            buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAll)
            
    regression_data = classification_low
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, classification_high['mlpValue'], classification_high['kNeighboursValue'])
        sell_pattern_without_mlalgo(regression_data, regressionResult, ws_buyPattern2, ws_sellPattern2)
        sell_oi_candidate(regression_data, regressionResult, None)
        all_withoutml(regression_data, regressionResult, ws_sellOI)
        sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, ws_sellPattern2)
        if sell_all_rule(regression_data, regressionResult, sellIndiaAvg, None):
            sell_year_high(regression_data, regressionResult, ws_sellYearHigh, ws_sellYearHigh1)
            sell_year_low(regression_data, regressionResult, ws_sellYearLow)
            sell_final(regression_data, regressionResult, ws_sellFinal, ws_sellFinal1)
            sell_high_indicators(regression_data, regressionResult, ws_sellHighIndicators)
            sell_up_trend(regression_data, regressionResult, None)
            sell_down_trend(regression_data, regressionResult, None)
            sell_oi(regression_data, regressionResult, None)
            sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAll)                                 
                          
def calculateParallel(threads=2, futures=None):
    pool = ThreadPool(threads)
    scrips = []
    for data in db.scrip.find({'futures':futures}):
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    pool.map(result_data, scrips)       
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1])
    connection.close()
    saveReports(sys.argv[1])